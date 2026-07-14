from openpyxl import load_workbook

from app.database import FILE_PATH
from helpers import customer_rows, dismiss_messagebox, unique_customer


def _assert_customer_exists(customer):
    rows = customer_rows()
    assert ("CustomerName", "CIN", "Email") in rows
    assert (customer["name"], customer["cin"], customer["email"]) in rows


def test_create_customer_success(customer_page):
    customer = unique_customer(1)

    customer_page.add_customer(customer["name"], customer["cin"], customer["email"])
    dismiss_messagebox("Success")

    _assert_customer_exists(customer)


def test_create_customer_trims_spaces(customer_page):
    customer = unique_customer(2)

    customer_page.add_customer(
        f"  {customer['name']}  ",
        f"  {customer['cin']}  ",
        f"  {customer['email']}  ",
    )
    dismiss_messagebox("Success")

    workbook = load_workbook(FILE_PATH)
    rows = list(workbook.active.iter_rows(values_only=True))
    assert (customer["name"], customer["cin"], customer["email"]) in rows


def test_create_customer_duplicate_cin_shows_error(customer_page):
    customer = unique_customer(3)

    customer_page.add_customer(customer["name"], customer["cin"], customer["email"])
    dismiss_messagebox("Success")

    customer_page.add_customer("Other Customer", customer["cin"], "other@example.com")
    dismiss_messagebox("Error")


def test_create_customer_missing_name_shows_error(customer_page):
    customer = unique_customer(4)

    customer_page.add_customer("", customer["cin"], customer["email"])
    dismiss_messagebox("Error")


def test_create_customer_invalid_cin_shows_error(customer_page):
    customer = unique_customer(5)

    customer_page.add_customer(customer["name"], "123", customer["email"])
    dismiss_messagebox("Error")


def test_create_customer_invalid_email_shows_error(customer_page):
    customer = unique_customer(6)

    customer_page.add_customer(customer["name"], customer["cin"], "invalid-email")
    dismiss_messagebox("Error")


def test_create_customer_missing_email_shows_error(customer_page):
    customer = unique_customer(19)

    customer_page.add_customer(customer["name"], customer["cin"], "")
    dismiss_messagebox("Error")
