from pathlib import Path
import tempfile

import time

from openpyxl import Workbook, load_workbook
from pywinauto import Desktop

from app.database import FILE_PATH, init_db


def reset_test_customer_database_deprecated():
    """Réinitialise le fichier runtime utilisé par les tests.

    - Source : data/customers.xlsx (manuel, modifiable)
    - Cible  : data/test_customers_runtime.xlsx (données pendant l'exécution)
    """
    src = Path(__file__).resolve().parent / "data" / "customers.xlsx"
    dst = Path(__file__).resolve().parent / "data" / "test_customers_runtime.xlsx"

    wb_src = load_workbook(src)
    wb_dst = Workbook()
    ws_dst = wb_dst.active

    # Recopie la feuille active complète (valeurs uniquement)
    ws_src = wb_src.active
    ws_dst.title = ws_src.title
    for row in ws_src.iter_rows(values_only=True):
        ws_dst.append(list(row))

    wb_dst.save(dst)


# Gardé pour compatibilité (peut être utilisé ailleurs)
def reset_customer_database():
    init_db()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["CustomerName", "CIN", "Email"])

    target_path = Path(FILE_PATH)
    # Windows peut refuser l'écriture si le fichier est verrouillé.
    # On tente plusieurs fois avant d'abandonner.
    if target_path.exists():
        try:
            target_path.unlink()
        except Exception:
            pass

    last_exc = None
    for _ in range(10):
        try:
            workbook.save(target_path)
            return
        except PermissionError as e:
            last_exc = e
            time.sleep(0.2)

    if last_exc is not None:
        raise last_exc




def customer_rows():
    workbook = load_workbook(FILE_PATH)
    worksheet = workbook.active
    return list(worksheet.iter_rows(values_only=True))


def save_customers(rows):
    workbook = load_workbook(FILE_PATH)
    worksheet = workbook.active

    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)

    for row in rows:
        worksheet.append(row)

    workbook.save(FILE_PATH)


def wait_for_window(title, timeout=10):
    desktop = Desktop(backend="win32")
    deadline = time.time() + timeout

    while time.time() < deadline:
        windows = desktop.windows(title=title)
        if windows:
            return windows[-1]
        time.sleep(0.2)

    raise RuntimeError(f"{title} window not found")


def dismiss_messagebox(title, timeout=5):
    """Ferme une messagebox Tk/Win32.

    Parfois pywinauto ne voit pas la fenêtre exactement avec le même titre.
    On tente donc plusieurs stratégies avec un retry.
    """
    last_exc = None
    for _ in range(10):
        try:
            dialog = wait_for_window(title, timeout=timeout)
            try:
                dialog.child_window(title="OK", class_name="Button").click_input()
            except Exception:
                try:
                    dialog.close()
                except Exception:
                    pass
            return dialog
        except Exception as e:
            last_exc = e
            time.sleep(0.2)

    # Fallback: si la fenêtre n'apparaît pas avec le titre, on ne casse pas.
    if last_exc is not None:
        raise last_exc



def unique_customer(index: int):

    """Génère un client de test.

    Objectif: utiliser des valeurs plus réalistes que "Test Customer N".
    """
    # Dataset proche de la réalité (noms FR/Ar) pour éviter l'impression de données fictives.
    # On le charge depuis un fichier pour pouvoir l'étendre facilement.
    # Source dataset : data/customers.xlsx (dataset utilisé par les tests)
    dataset_path = Path(__file__).resolve().parent / "data" / "customers.xlsx"

    try:
        from openpyxl import load_workbook

        wb = load_workbook(dataset_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        # expected header: CustomerName | CIN | Email
        dataset = []
        for r in rows[1:]:
            if not r or not r[0] or not r[1] or not r[2]:
                continue
            dataset.append({"name": str(r[0]).strip(), "cin": str(r[1]).strip(), "email": str(r[2]).strip()})

        # Si le fichier ne contient aucun client (max_row=1), on remplit avec un fallback.
        if not dataset:
            raise ValueError("No customer rows found in data/customers.xlsx")


    except Exception:

        # fallback en dur (au cas où le fichier fixtures serait indisponible)
        dataset = [
            {"name": "Amine Ben Salah", "cin": "91000001", "email": "amine.bensalah@example.com"},
            {"name": "Sara Trabelsi", "cin": "91000002", "email": "sara.trabelsi@example.com"},
            {"name": "Karim Mokhtar", "cin": "91000003", "email": "karim.mokhtar@example.com"},
            {"name": "Nour El Amri", "cin": "91000004", "email": "nour.elamri@example.com"},
            {"name": "Hedi Jarray", "cin": "91000005", "email": "hedi.jarray@example.com"},
            {"name": "Salma Ksouri", "cin": "91000006", "email": "salma.ksouri@example.com"},
            {"name": "Yassine Ayadi", "cin": "91000007", "email": "yassine.ayadi@example.com"},
            {"name": "Ines Gharbi", "cin": "91000008", "email": "ines.gharbi@example.com"},
            {"name": "Omar Cherif", "cin": "91000009", "email": "omar.cherif@example.com"},
            {"name": "Mariam Ben Romdhane", "cin": "91000010", "email": "mariam.benromdhane@example.com"},
            {"name": "Ibrahim Saadi", "cin": "91000011", "email": "ibrahim.saadi@example.com"},
            {"name": "Lina Ben Ammar", "cin": "91000012", "email": "lina.benammar@example.com"},
        ]


    # Les tests passent index (ex: 1,2,3...). On mappe sur le dataset en faisant -1.
    idx = max(index - 1, 0)
    if idx < len(dataset):
        return dataset[idx]

    # Si l'index demandé n'existe pas dans data/customers.xlsx, on génère un client fictif
    # mais valide (format réaliste) pour que les tests puissent continuer.
    # CIN: 8 chiffres commençant par 91 + index (padding)
    cin = f"91{index:06d}"
    name = f"Test Customer {index}"
    email = f"customer{index}@example.com"
    return {"name": name, "cin": cin, "email": email}



