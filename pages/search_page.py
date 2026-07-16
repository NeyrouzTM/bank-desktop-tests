import time

from openpyxl import load_workbook

from app.database import FILE_PATH
from pages.base_page import BaseTkPage
from pages.dashboard_page import DashboardPage


class SearchPage(BaseTkPage):

    window_title = "Dashboard"

    def __init__(self):
        super().__init__()
        dashboard = DashboardPage()
        dashboard.open_search_window()
        self._attach_window(self.window_title)

        # Wait until the CIN input exists (timing is flaky on Tk redraw)
        entry_controls = []
        deadline = time.time() + 10
        while time.time() < deadline:
            entry_controls = self._entry_controls()
            if entry_controls:
                break
            time.sleep(0.2)

        if not entry_controls:
            raise RuntimeError("Search CIN field not found")


        self.cin_field = entry_controls[0]
        self.result_text = ""
        self.search_button = self._button_by_text("Search")

    def search_customer(self, cin):
        lookup_cin = cin.strip()

        self._click_and_type(self.cin_field, cin)
        self.search_button.click_input()
        time.sleep(0.5)

        # Runtime Excel may not be flushed yet when Tk triggers the search.
        # Retry to avoid BadZipFile / partially written XLSX.
        last_exc = None
        workbook = None
        for _ in range(10):
            try:
                workbook = load_workbook(FILE_PATH)
                break
            except Exception as e:
                last_exc = e
                time.sleep(0.2)

        if workbook is None:
            raise last_exc

        worksheet = workbook.active


        for row in worksheet.iter_rows(values_only=True):
            if row[1] == lookup_cin:
                self.result_text = f"Name: {row[0]} | CIN: {row[1]} | Email: {row[2]}"
                return self.result_text

        self.result_text = "Customer not found"
        return self.result_text
