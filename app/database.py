from openpyxl import Workbook, load_workbook
import os
from datetime import datetime, timedelta
import hashlib
import random

FILE_PATH = "data/customers.xlsx"
TRANSFERS_PATH = "data/transfers.xlsx"
ACTIVITY_PATH = "data/activity.xlsx"

TRANSFER_HEADERS = [
    "Date",
    "From",
    "To",
    "Amount",
    "Status",
    "Region",
    "Channel",
    "Country",
    "TransferId",
]

ACTIVITY_HEADERS = [
    "Date",
    "EventType",
    "Reference",
    "From",
    "To",
    "Amount",
    "Status",
    "Region",
    "Channel",
    "Country",
    "Detail",
]

REGION_META = {
    "Tunis": "Tunisia",
    "Sfax": "Tunisia",
    "Sousse": "Tunisia",
    "Paris": "France",
    "Lyon": "France",
    "Marseille": "France",
}

REGIONS = list(REGION_META.keys())
CHANNELS = ["Mobile App", "Web Banking", "Branch", "ATM", "Desktop App"]
STATUSES = ["Success", "Pending", "Failed"]


def init_db():
    if not os.path.exists("data"):
        os.makedirs("data")

    if not os.path.exists(FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(["CustomerName", "CIN", "Email"])
        wb.save(FILE_PATH)

    init_transfers_db()
    init_activity_db()


def init_transfers_db(seed_if_empty: bool = True):
    if not os.path.exists("data"):
        os.makedirs("data")

    created = False
    if not os.path.exists(TRANSFERS_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(TRANSFER_HEADERS)
        wb.save(TRANSFERS_PATH)
        created = True
    else:
        wb = load_workbook(TRANSFERS_PATH)
        ws = wb.active

    if seed_if_empty and (created or ws.max_row <= 1):
        _seed_transfers()


def init_activity_db():
    if not os.path.exists("data"):
        os.makedirs("data")

    if not os.path.exists(ACTIVITY_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(ACTIVITY_HEADERS)
        wb.save(ACTIVITY_PATH)


def safe_load():
    if not os.path.exists(FILE_PATH):
        init_db()
    return load_workbook(FILE_PATH)


def safe_load_transfers():
    init_transfers_db(seed_if_empty=False)
    return load_workbook(TRANSFERS_PATH)


def safe_load_activity():
    init_activity_db()
    return load_workbook(ACTIVITY_PATH)


def add_customer(name, cin, email):
    wb = safe_load()
    ws = wb.active

    for row in ws.iter_rows(values_only=True):
        if row[1] == cin:
            return False

    ws.append([name, cin, email])
    wb.save(FILE_PATH)

    region = _pick_region(cin)
    log_activity(
        event_type="CustomerCreated",
        reference=cin,
        from_cin="",
        to_cin=cin,
        amount=0,
        status="Success",
        region=region,
        channel="Desktop App",
        detail=f"Customer {name} created",
    )
    return True


def _pick_region(key: str) -> str:
    digest = hashlib.md5(str(key).encode("utf-8")).hexdigest()
    return REGIONS[int(digest, 16) % len(REGIONS)]


def _pick_channel(key: str) -> str:
    digest = hashlib.md5(f"channel-{key}".encode("utf-8")).hexdigest()
    return CHANNELS[int(digest, 16) % len(CHANNELS)]


def _next_transfer_id(ws) -> str:
    count = max(ws.max_row - 1, 0) + 1
    return f"TRX-{100000 + count}"


def add_transfer(
    from_cin: str,
    to_cin: str,
    amount: float,
    status: str = "Success",
    region: str | None = None,
    channel: str | None = None,
    when: datetime | None = None,
):
    wb = safe_load_transfers()
    ws = wb.active

    region = region or _pick_region(from_cin)
    channel = channel or _pick_channel(f"{from_cin}-{to_cin}-{amount}")
    country = REGION_META.get(region, "Tunisia")
    when = when or datetime.now()
    transfer_id = _next_transfer_id(ws)
    date_str = when.strftime("%Y-%m-%d %H:%M:%S")

    ws.append(
        [
            date_str,
            from_cin,
            to_cin,
            float(amount),
            status,
            region,
            channel,
            country,
            transfer_id,
        ]
    )
    wb.save(TRANSFERS_PATH)

    log_activity(
        event_type="Transfer",
        reference=transfer_id,
        from_cin=from_cin,
        to_cin=to_cin,
        amount=amount,
        status=status,
        region=region,
        channel=channel,
        detail=f"Transfer {from_cin} → {to_cin}",
        when=when,
    )
    return transfer_id


def log_activity(
    event_type: str,
    reference: str,
    from_cin: str,
    to_cin: str,
    amount: float,
    status: str,
    region: str,
    channel: str,
    detail: str,
    when: datetime | None = None,
):
    wb = safe_load_activity()
    ws = wb.active
    when = when or datetime.now()
    country = REGION_META.get(region, "Tunisia")
    ws.append(
        [
            when.strftime("%Y-%m-%d %H:%M:%S"),
            event_type,
            reference,
            from_cin,
            to_cin,
            float(amount or 0),
            status,
            region,
            channel,
            country,
            detail,
        ]
    )
    wb.save(ACTIVITY_PATH)


def list_transfers() -> list[dict]:
    wb = safe_load_transfers()
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        rows.append(
            {
                "date": str(row[0]),
                "from": str(row[1] or ""),
                "to": str(row[2] or ""),
                "amount": float(row[3] or 0),
                "status": str(row[4] or "Success"),
                "region": str(row[5] or ""),
                "channel": str(row[6] or ""),
                "country": str(row[7] or REGION_META.get(str(row[5] or ""), "")),
                "transfer_id": str(row[8] or ""),
            }
        )
    return rows


def list_activity() -> list[dict]:
    wb = safe_load_activity()
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        rows.append(
            {
                "date": str(row[0]),
                "event_type": str(row[1] or ""),
                "reference": str(row[2] or ""),
                "from": str(row[3] or ""),
                "to": str(row[4] or ""),
                "amount": float(row[5] or 0),
                "status": str(row[6] or ""),
                "region": str(row[7] or ""),
                "channel": str(row[8] or ""),
                "country": str(row[9] or ""),
                "detail": str(row[10] or ""),
            }
        )
    return rows


def customer_count() -> int:
    try:
        wb = safe_load()
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        return sum(1 for row in rows if row and row[0])
    except Exception:
        return 0


def _seed_transfers():
    """Starter history so the BI is not empty on first launch."""
    rng = random.Random(42)
    now = datetime.now()
    samples = [
        ("10000001", "10000002", 450.0),
        ("10000003", "10000004", 1200.5),
        ("10000005", "10000001", 89.9),
        ("10000002", "10000006", 3200.0),
        ("10000007", "10000003", 760.25),
        ("10000004", "10000008", 15000.0),
        ("10000009", "10000005", 210.0),
        ("10000001", "10000009", 980.0),
        ("10000006", "10000007", 5400.0),
        ("10000008", "10000002", 175.4),
        ("10000003", "10000001", 2499.99),
        ("10000005", "10000004", 65.0),
    ]
    for index, (src, dst, amount) in enumerate(samples):
        days_ago = rng.randint(0, 100)
        when = now - timedelta(days=days_ago, hours=rng.randint(0, 20))
        status = STATUSES[0] if index % 7 else (STATUSES[1] if index % 5 else STATUSES[2])
        if index % 7 not in (0, 1):
            status = "Success"
        elif index % 7 == 1:
            status = "Pending"
        else:
            status = "Failed" if index % 11 == 0 else "Success"
        region = REGIONS[index % len(REGIONS)]
        channel = CHANNELS[index % len(CHANNELS)]
        add_transfer(
            src,
            dst,
            amount,
            status=status,
            region=region,
            channel=channel,
            when=when,
        )


# Ensure storage exists when the module is imported by the app.
init_db()
