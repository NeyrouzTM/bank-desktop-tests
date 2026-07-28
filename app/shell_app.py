"""Single-window banking shell with premium Vermeg UX."""

from __future__ import annotations

import re
import tkinter as tk
from tkinter import messagebox

from openpyxl import load_workbook

from database import FILE_PATH, add_customer, add_insurance_subscription, add_transfer, customer_count, list_insurance_subscriptions, list_transfers
from theme import DARK, LIGHT, FONT_BODY, FONT_BUTTON, FONT_SMALL, FONT_SUBTITLE, FONT_TITLE
from widgets import (
    expose_win_text,
    fade_in,
    glass_card,
    inline_alert,
    primary_button,
    set_alert,
    styled_entry,
    styled_label,
    vermeg_logo,
)
from ui_background import apply_window_background


NAV_ITEMS = [
    ("Dashboard", "dashboard"),
    ("Create Customer", "customers"),
    ("Search Customer", "search"),
    ("Transfer Money", "transfer"),
    ("Insurance", "insurance"),
    ("Power BI Dashboard", "bi"),
]


class BankApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Login")
        self.root.geometry("1180x760")
        self.root.minsize(1020, 680)
        self.theme_name = "dark"
        self.theme = dict(DARK)
        self.current_view = None
        self.nav_buttons: dict[str, tk.Button] = {}
        self._views: dict[str, tk.Frame] = {}

        apply_window_background(self.root)
        self._build_login()

    @property
    def t(self):
        return self.theme

    def run(self):
        self.root.mainloop()

    def toggle_theme(self):
        self.theme_name = "light" if self.theme_name == "dark" else "dark"
        self.theme = dict(LIGHT if self.theme_name == "light" else DARK)
        # Full rebuild of visible stage keeps styling consistent
        if self.root.title() == "Login":
            self._rebuild_login()
        else:
            self._rebuild_shell(self.current_view or "dashboard")

    # ---------------- LOGIN ----------------
    def _build_login(self):
        self._clear_root()
        self.root.title("Login")
        t = self.t

        overlay = tk.Frame(self.root, bg=t["bg"])
        overlay.place(relx=0.5, rely=0.5, anchor="center")

        card_host = tk.Frame(overlay, bg=t["card_border"])
        card_host.pack()
        card = tk.Frame(card_host, bg=t["card"])
        card.pack(padx=1, pady=1)

        body = tk.Frame(card, bg=t["card"])
        body.pack(padx=36, pady=32)

        logo_bg = dict(t)
        logo_bg["logo_bg"] = t["card"]
        vermeg_logo(body, logo_bg, size=26).pack(anchor="w")

        styled_label(
            body,
            t,
            "Banking Workspace",
            font=FONT_TITLE,
            fg=t["text"],
            bg=t["card"],
        ).pack(anchor="w", pady=(14, 2))

        styled_label(
            body,
            t,
            "Sign in to access customers, transfers and live BI.",
            font=FONT_SUBTITLE,
            fg=t["muted"],
            bg=t["card"],
        ).pack(anchor="w", pady=(0, 18))

        styled_label(body, t, "Username", font=FONT_SMALL, fg=t["muted"], bg=t["card"]).pack(
            anchor="w"
        )
        # Placeholder labels keep layout; real Entry/Button are direct children of root
        # so pywinauto (win32) can find them reliably.
        tk.Frame(body, bg=t["card"], height=28).pack(anchor="w", pady=(4, 10), fill="x")

        styled_label(body, t, "Password", font=FONT_SMALL, fg=t["muted"], bg=t["card"]).pack(
            anchor="w"
        )
        tk.Frame(body, bg=t["card"], height=28).pack(anchor="w", pady=(4, 8), fill="x")

        self.login_alert = inline_alert(body, t)
        self.login_alert.pack(anchor="w", fill="x", pady=(4, 10), ipady=6)

        actions = tk.Frame(body, bg=t["card"], height=40)
        actions.pack(fill="x", pady=(4, 0))
        actions.pack_propagate(False)

        theme_lbl = tk.Label(
            actions,
            text="Dark / Light",
            bg=t["card"],
            fg=t["cyan"],
            font=FONT_SMALL,
            cursor="hand2",
        )
        theme_lbl.pack(side="right", padx=(12, 0), pady=8)
        theme_lbl.bind("<Button-1>", lambda _e: self.toggle_theme())

        hint = styled_label(
            body,
            t,
            "Demo  ·  admin / admin123",
            font=FONT_SMALL,
            fg=t["muted"],
            bg=t["card"],
        )
        hint.pack(anchor="w", pady=(16, 0))

        # Automation-friendly controls: direct children of the Tk root
        self.root.update_idletasks()
        self.login_user = styled_entry(self.root, t, width=34)
        self.login_pass = styled_entry(self.root, t, show="*", width=34)
        self.login_button = primary_button(self.root, t, "Login", self._do_login, width=16)

        def _place_login_controls(_event=None):
            try:
                cx = self.root.winfo_width() // 2
                cy = self.root.winfo_height() // 2
                self.login_user.place(x=cx - 140, y=cy - 10, width=280, height=28)
                self.login_pass.place(x=cx - 140, y=cy + 48, width=280, height=28)
                self.login_button.place(x=cx - 140, y=cy + 100, width=160, height=36)
                self.login_user.lift()
                self.login_pass.lift()
                self.login_button.lift()
                expose_win_text(self.login_button, "Login")
            except tk.TclError:
                pass

        self.root.bind("<Configure>", _place_login_controls, add="+")
        self.root.after(50, _place_login_controls)

        self.login_user.bind("<Return>", lambda _e: self.login_pass.focus())
        self.login_pass.bind("<Return>", lambda _e: self._do_login())
        self.login_user.focus()
        fade_in(card_host)

    def _rebuild_login(self):
        self._build_login()

    def _do_login(self):
        username = self.login_user.get().strip()
        password = self.login_pass.get().strip()
        t = self.t

        if not username:
            set_alert(self.login_alert, t, "Username is required.")
            self.login_user.focus()
            return
        if not password:
            set_alert(self.login_alert, t, "Password is required.")
            self.login_pass.focus()
            return
        if username != "admin" or password != "admin123":
            set_alert(self.login_alert, t, "Invalid credentials. Please try again.")
            self.login_pass.delete(0, tk.END)
            self.login_pass.focus()
            return

        set_alert(self.login_alert, t, "Welcome — opening workspace…", ok=True)
        self._rebuild_shell("dashboard")

    # ---------------- SHELL ----------------
    def _rebuild_shell(self, view_key: str):
        self._clear_root()
        self.root.title("Dashboard")
        t = self.t

        shell = tk.Frame(self.root, bg=t["bg"])
        shell.pack(fill="both", expand=True)

        # Sidebar
        sidebar = tk.Frame(shell, bg=t["sidebar"], width=220)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        brand = tk.Frame(sidebar, bg=t["sidebar"])
        brand.pack(fill="x", padx=18, pady=(22, 10))
        logo_theme = dict(t)
        logo_theme["logo_bg"] = t["sidebar"]
        vermeg_logo(brand, logo_theme, size=18).pack(anchor="w")
        styled_label(
            brand,
            t,
            "Bank Desktop",
            font=FONT_SMALL,
            fg=t["muted"],
            bg=t["sidebar"],
        ).pack(anchor="w", pady=(6, 0))

        nav_wrap = tk.Frame(sidebar, bg=t["sidebar"])
        nav_wrap.pack(fill="x", padx=12, pady=12)

        # Visual placeholders in sidebar (real Buttons are placed on root for pywinauto)
        self.nav_buttons = {}
        for label, key in NAV_ITEMS:
            spacer = tk.Frame(nav_wrap, bg=t["sidebar"], height=42)
            spacer.pack(fill="x", pady=3)

        footer = tk.Frame(sidebar, bg=t["sidebar"])
        footer.pack(side="bottom", fill="x", padx=12, pady=16)

        theme_chip = tk.Label(
            footer,
            text=f"{'Light' if self.theme_name == 'dark' else 'Dark'} mode",
            bg=t["sidebar"],
            fg=t["cyan"],
            font=FONT_SMALL,
            cursor="hand2",
        )
        theme_chip.pack(anchor="w", padx=8, pady=8)
        theme_chip.bind("<Button-1>", lambda _e: self.toggle_theme())

        # Main column
        main = tk.Frame(shell, bg=t["bg"])
        main.pack(side="left", fill="both", expand=True)

        topbar = tk.Frame(main, bg=t["bg_alt"], height=56)
        topbar.pack(fill="x")
        topbar.pack_propagate(False)

        self.topbar_title = styled_label(
            topbar, t, "Dashboard", font=FONT_TITLE, fg=t["text"], bg=t["bg_alt"]
        )
        self.topbar_title.pack(side="left", padx=22, pady=12)

        styled_label(
            topbar,
            t,
            "admin  ·  Vermeg Banking",
            font=FONT_SMALL,
            fg=t["muted"],
            bg=t["bg_alt"],
        ).pack(side="right", padx=22)

        self.content = tk.Frame(main, bg=t["bg"])
        self.content.pack(fill="both", expand=True, padx=18, pady=16)

        self._views = {}
        self._mount_nav_buttons()
        self.show_view(view_key)

    def _mount_nav_buttons(self):
        """Nav Buttons as direct children of root — required for pywinauto win32."""
        t = self.t
        self.nav_buttons = {}
        for index, (label, key) in enumerate(NAV_ITEMS):
            btn = tk.Button(
                self.root,
                text=label,
                anchor="w",
                command=lambda k=key: self.show_view(k),
                bg=t["sidebar"],
                fg=t["text"],
                activebackground=t["hover"],
                activeforeground=t["text"],
                relief="flat",
                bd=0,
                padx=14,
                font=FONT_BUTTON,
                cursor="hand2",
            )
            btn.place(x=16, y=100 + index * 48, width=190, height=40)
            btn.lift()
            self.root.update_idletasks()
            expose_win_text(btn, label)
            self._bind_nav_hover(btn)
            self.nav_buttons[key] = btn

    def _clear_form_controls(self):
        for attr in (
            "form_entries",
            "form_button",
            "_form_name",
            "_form_cin",
            "_form_email",
            "_form_from",
            "_form_to",
            "_form_amount",
            "_form_search_cin",
            "_form_insurance_cin",
            "_form_insurance_balance",
            "_insurance_search_btn",
            "_insurance_pay_btn",
        ):
            widget = getattr(self, attr, None)
            if widget is None:
                continue
            if isinstance(widget, (list, tuple)):
                for item in widget:
                    try:
                        item.destroy()
                    except tk.TclError:
                        pass
            else:
                try:
                    widget.destroy()
                except tk.TclError:
                    pass
            setattr(self, attr, None)

    def _bind_nav_hover(self, btn: tk.Button):
        t = self.t

        def enter(_e, b=btn):
            if b.cget("bg") != t["purple"]:
                b.configure(bg=t["hover"])

        def leave(_e, b=btn):
            if b.cget("bg") != t["purple"]:
                b.configure(bg=t["sidebar"])

        btn.bind("<Enter>", enter)
        btn.bind("<Leave>", leave)

    def _highlight_nav(self, key: str):
        t = self.t
        for k, btn in self.nav_buttons.items():
            if k == key:
                btn.configure(bg=t["purple"], fg="#ffffff")
            else:
                btn.configure(bg=t["sidebar"], fg=t["text"])

    def show_view(self, key: str):
        titles = {
            "dashboard": "Dashboard",
            "customers": "Create Customer",
            "search": "Search Customer",
            "transfer": "Transfer Money",
            "insurance": "Insurance",
            "bi": "Power BI Dashboard",
        }
        self.current_view = key
        self._highlight_nav(key)
        self.topbar_title.configure(text=titles.get(key, "Dashboard"))
        self._clear_form_controls()

        for child in self.content.winfo_children():
            child.destroy()

        builders = {
            "dashboard": self._view_dashboard,
            "customers": self._view_customers,
            "search": self._view_search,
            "transfer": self._view_transfer,
            "insurance": self._view_insurance,
            "bi": self._view_bi,
        }
        builders[key]()
        fade_in(self.content)
        # Keep nav buttons above content
        for btn in self.nav_buttons.values():
            try:
                btn.lift()
            except tk.TclError:
                pass

    def _clear_root(self):
        self._clear_form_controls()
        try:
            self.root.unbind("<Configure>")
        except tk.TclError:
            pass
        for child in self.root.winfo_children():
            child.destroy()
        # Re-apply background after wipe
        apply_window_background(self.root)

    def _place_on_slot(self, widget: tk.Widget, slot: tk.Widget):
        self.root.update_idletasks()
        widget.place(
            x=slot.winfo_rootx() - self.root.winfo_rootx(),
            y=slot.winfo_rooty() - self.root.winfo_rooty(),
            width=slot.winfo_width(),
            height=slot.winfo_height(),
        )
        widget.lift()

    # ---------------- VIEWS ----------------
    def _view_dashboard(self):
        t = self.t
        wrap = tk.Frame(self.content, bg=t["bg"])
        wrap.pack(fill="both", expand=True)

        hero = glass_card(wrap, t)
        hero.master.master.pack(fill="x", pady=(0, 14))

        styled_label(hero, t, "Welcome back", font=FONT_SMALL, fg=t["muted"]).pack(anchor="w")
        styled_label(
            hero,
            t,
            "Vermeg Banking Control Center",
            font=("Segoe UI Semibold", 20),
        ).pack(anchor="w", pady=(4, 6))
        styled_label(
            hero,
            t,
            "Manage customers, simulate transfers and monitor live Power BI analytics — all in one workspace.",
            font=FONT_BODY,
            fg=t["muted"],
            wraplength=760,
            justify="left",
        ).pack(anchor="w")

        stats = tk.Frame(wrap, bg=t["bg"])
        stats.pack(fill="x", pady=(0, 14))

        transfers = list_transfers()
        volume = sum(x["amount"] for x in transfers)
        success = sum(1 for x in transfers if x["status"] == "Success")
        rate = round((success / len(transfers)) * 100, 1) if transfers else 0

        cards = [
            ("Customers", str(customer_count()), t["cyan"]),
            ("Transfers", str(len(transfers)), t["purple"]),
            ("Volume", f"{volume:,.0f}", t["amber"]),
            ("Success rate", f"{rate}%", t["green"]),
        ]
        for i, (title, value, color) in enumerate(cards):
            cell = glass_card(stats, t, padx=14, pady=12)
            cell.master.master.grid(row=0, column=i, sticky="nsew", padx=(0 if i == 0 else 10, 0))
            stats.grid_columnconfigure(i, weight=1)
            tk.Frame(cell, bg=color, height=3).pack(fill="x", pady=(0, 8))
            styled_label(cell, t, title.upper(), font=FONT_SMALL, fg=t["muted"]).pack(anchor="w")
            styled_label(cell, t, value, font=("Segoe UI Semibold", 18)).pack(anchor="w", pady=(4, 0))

        actions = glass_card(wrap, t)
        actions.master.master.pack(fill="both", expand=True)
        styled_label(actions, t, "Quick actions", font=("Segoe UI Semibold", 12)).pack(anchor="w")
        row = tk.Frame(actions, bg=t["card"])
        row.pack(anchor="w", pady=12)

        shortcuts = [
            ("Create Customer", "customers"),
            ("Search Customer", "search"),
            ("Transfer Money", "transfer"),
            ("Power BI Dashboard", "bi"),
        ]
        for text, key in shortcuts:
            # Use Labels (not Buttons) so automation keeps a single nav Button per destination.
            chip = tk.Label(
                row,
                text=text,
                bg=t["purple"],
                fg="#ffffff",
                padx=14,
                pady=8,
                font=FONT_BUTTON,
                cursor="hand2",
            )
            chip.pack(side="left", padx=(0, 10))
            chip.bind("<Button-1>", lambda _e, k=key: self.show_view(k))
            chip.bind("<Enter>", lambda _e, c=chip: c.configure(bg=t["accent"]))
            chip.bind("<Leave>", lambda _e, c=chip: c.configure(bg=t["purple"]))

    def _view_customers(self):
        t = self.t
        email_pattern = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")

        panel = glass_card(self.content, t)
        panel.master.master.pack(fill="both", expand=True)

        styled_label(panel, t, "Create Customer", font=FONT_TITLE).pack(anchor="w")
        styled_label(
            panel, t, "Onboard a new banking customer", font=FONT_SUBTITLE, fg=t["muted"]
        ).pack(anchor="w", pady=(2, 16))

        form = tk.Frame(panel, bg=t["card"])
        form.pack(anchor="w")

        entry_slots = []
        for field in ("Customer Name", "CIN", "Email"):
            styled_label(form, t, field, font=FONT_SMALL, fg=t["muted"]).pack(anchor="w")
            slot = tk.Frame(form, bg=t["card"], width=320, height=28)
            slot.pack(anchor="w", pady=(6, 12))
            slot.pack_propagate(False)
            entry_slots.append(slot)

        alert = inline_alert(form, t)
        alert.pack(anchor="w", fill="x", pady=(0, 10), ipady=5)
        button_slot = tk.Frame(form, bg=t["card"], width=180, height=36)
        button_slot.pack(anchor="w")
        button_slot.pack_propagate(False)

        def save():
            name = self._form_name.get().strip()
            cin = self._form_cin.get().strip()
            email = self._form_email.get().strip()

            if not name or not cin or not email:
                messagebox.showerror("Error", "All fields are required")
                return
            if len(name) < 3:
                messagebox.showerror("Error", "Name must contain at least 3 characters")
                return
            if not cin.isdigit() or len(cin) != 8:
                messagebox.showerror("Error", "Invalid CIN")
                return
            if not email_pattern.fullmatch(email):
                messagebox.showerror("Error", "Invalid email format")
                return

            success = add_customer(name, cin, email)
            if success:
                messagebox.showinfo("Success", "Customer added successfully")
                self._form_name.delete(0, tk.END)
                self._form_cin.delete(0, tk.END)
                self._form_email.delete(0, tk.END)
                set_alert(alert, t, "Customer saved and sent to live BI history.", ok=True)
                self._form_name.focus()
            else:
                messagebox.showerror("Error", "CIN already exists")

        # Direct-on-root controls for automation
        self._form_name = styled_entry(self.root, t)
        self._form_cin = styled_entry(self.root, t)
        self._form_email = styled_entry(self.root, t)
        self.form_button = primary_button(self.root, t, "Save Customer", save, width=20)
        self.form_entries = [self._form_name, self._form_cin, self._form_email]

        def place_form(_e=None):
            try:
                for entry, slot in zip(self.form_entries, entry_slots):
                    self._place_on_slot(entry, slot)
                self._place_on_slot(self.form_button, button_slot)
                for btn in self.nav_buttons.values():
                    btn.lift()
            except tk.TclError:
                pass

        self.root.after(40, place_form)

    def _view_search(self):
        t = self.t
        panel = glass_card(self.content, t)
        panel.master.master.pack(fill="both", expand=True)

        styled_label(panel, t, "Search Customer", font=FONT_TITLE).pack(anchor="w")
        styled_label(
            panel, t, "Lookup by CIN in the customer ledger", font=FONT_SUBTITLE, fg=t["muted"]
        ).pack(anchor="w", pady=(2, 16))

        styled_label(panel, t, "Enter CIN", font=FONT_SMALL, fg=t["muted"]).pack(anchor="w")
        entry_slot = tk.Frame(panel, bg=t["card"], width=320, height=28)
        entry_slot.pack(anchor="w", pady=(6, 12))
        entry_slot.pack_propagate(False)

        result = styled_label(panel, t, "", font=FONT_BODY)
        result.pack(anchor="w", pady=(0, 12))
        button_slot = tk.Frame(panel, bg=t["card"], width=140, height=36)
        button_slot.pack(anchor="w")
        button_slot.pack_propagate(False)

        def search():
            cin = self._form_search_cin.get().strip()
            if not cin:
                messagebox.showerror("Error", "CIN required")
                return
            if not cin.isdigit() or len(cin) != 8:
                messagebox.showerror("Error", "Invalid CIN")
                return

            wb = load_workbook(FILE_PATH)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if row[1] == cin:
                    result.configure(
                        text=f"Name: {row[0]} | CIN: {row[1]} | Email: {row[2]}",
                        fg=t["green"],
                    )
                    return
            result.configure(text="Customer not found", fg=t["danger"])

        self._form_search_cin = styled_entry(self.root, t)
        self.form_button = primary_button(self.root, t, "Search", search, width=16)
        self.form_entries = [self._form_search_cin]

        def place_form(_e=None):
            try:
                self._place_on_slot(self._form_search_cin, entry_slot)
                self._place_on_slot(self.form_button, button_slot)
                for btn in self.nav_buttons.values():
                    btn.lift()
            except tk.TclError:
                pass

        self.root.after(40, place_form)

    def _view_transfer(self):
        t = self.t
        amount_pattern = re.compile(r"^\d+(\.\d{1,2})?$")

        panel = glass_card(self.content, t)
        panel.master.master.pack(fill="both", expand=True)

        styled_label(panel, t, "Transfer Money", font=FONT_TITLE).pack(anchor="w")
        styled_label(
            panel,
            t,
            "Simulate a transfer — each success feeds transfers.xlsx & BI",
            font=FONT_SUBTITLE,
            fg=t["muted"],
        ).pack(anchor="w", pady=(2, 16))

        entry_slots = []
        for field in ("From CIN", "To CIN", "Amount (simulation)"):
            styled_label(panel, t, field, font=FONT_SMALL, fg=t["muted"]).pack(anchor="w")
            slot = tk.Frame(panel, bg=t["card"], width=320, height=28)
            slot.pack(anchor="w", pady=(6, 12))
            slot.pack_propagate(False)
            entry_slots.append(slot)
        button_slot = tk.Frame(panel, bg=t["card"], width=140, height=36)
        button_slot.pack(anchor="w", pady=(8, 0))
        button_slot.pack_propagate(False)

        def transfer():
            from_cin = self._form_from.get().strip()
            to_cin = self._form_to.get().strip()
            amount = self._form_amount.get().strip().replace(",", ".")

            if not from_cin or not to_cin or not amount:
                messagebox.showerror("Error", "All fields required")
                return
            if from_cin == to_cin:
                messagebox.showerror("Error", "Cannot transfer to same account")
                return
            if not from_cin.isdigit() or len(from_cin) != 8:
                messagebox.showerror("Error", "Invalid sender CIN")
                return
            if not to_cin.isdigit() or len(to_cin) != 8:
                messagebox.showerror("Error", "Invalid recipient CIN")
                return
            if not amount_pattern.fullmatch(amount):
                messagebox.showerror("Error", "Amount format is invalid")
                return
            try:
                amount_value = float(amount)
                if amount_value <= 0:
                    messagebox.showerror("Error", "Invalid amount")
                    return
            except ValueError:
                messagebox.showerror("Error", "Amount must be numeric")
                return

            wb = load_workbook(FILE_PATH)
            ws = wb.active
            from_exists = False
            to_exists = False
            for row in ws.iter_rows(values_only=True):
                if row[1] == from_cin:
                    from_exists = True
                if row[1] == to_cin:
                    to_exists = True
            if not from_exists or not to_exists:
                messagebox.showerror("Error", "Invalid CIN(s)")
                return

            add_transfer(from_cin, to_cin, amount_value, status="Success")
            messagebox.showinfo("Success", "Transfer simulated successfully")
            self._form_from.delete(0, tk.END)
            self._form_to.delete(0, tk.END)
            self._form_amount.delete(0, tk.END)

        self._form_from = styled_entry(self.root, t)
        self._form_to = styled_entry(self.root, t)
        self._form_amount = styled_entry(self.root, t)
        self.form_button = primary_button(self.root, t, "Transfer", transfer, width=16)
        self.form_entries = [self._form_from, self._form_to, self._form_amount]

        def place_form(_e=None):
            try:
                for entry, slot in zip(self.form_entries, entry_slots):
                    self._place_on_slot(entry, slot)
                self._place_on_slot(self.form_button, button_slot)
                for btn in self.nav_buttons.values():
                    btn.lift()
            except tk.TclError:
                pass

        self.root.after(40, place_form)

    def _view_insurance(self):
        t = self.t

        # Pack catalog
        packs = {
            "hospital": [("Pack 1 - Hospitalisation de base", 200), ("Pack 2 - Hospitalisation confort", 500), ("Pack 3 - Hospitalisation premium", 800)],
            "ambulatoire": [("Pack 1 - Soins de base", 100), ("Pack 2 - Soins intermédiaires", 300), ("Pack 3 - Soins complets", 600)],
        }

        panel = glass_card(self.content, t)
        panel.master.master.pack(fill="both", expand=True)

        styled_label(panel, t, "Insurance", font=FONT_TITLE).pack(anchor="w")
        styled_label(
            panel, t, "Souscrivez à une assurance et choisissez votre pack", font=FONT_SUBTITLE, fg=t["muted"]
        ).pack(anchor="w", pady=(2, 16))

        # --- Customer selection ---
        customer_frame = tk.Frame(panel, bg=t["card"])
        customer_frame.pack(anchor="w", fill="x")

        styled_label(customer_frame, t, "Rechercher un client (CIN)", font=FONT_SMALL, fg=t["muted"]).pack(anchor="w")
        cin_slot = tk.Frame(customer_frame, bg=t["card"], width=200, height=28)
        cin_slot.pack(side="left", pady=(6, 12))
        cin_slot.pack_propagate(False)

        search_btn_slot = tk.Frame(customer_frame, bg=t["card"], width=100, height=28)
        search_btn_slot.pack(side="left", padx=(10, 0), pady=(6, 12))
        search_btn_slot.pack_propagate(False)

        self.customer_info_label = styled_label(panel, t, "", font=FONT_BODY, fg=t["green"])
        self.customer_info_label.pack(anchor="w", pady=(0, 8))

        # --- Balance ---
        balance_frame = tk.Frame(panel, bg=t["card"])
        balance_frame.pack(anchor="w", fill="x")

        styled_label(balance_frame, t, "Solde du compte (€)", font=FONT_SMALL, fg=t["muted"]).pack(anchor="w")
        balance_slot = tk.Frame(balance_frame, bg=t["card"], width=200, height=28)
        balance_slot.pack(anchor="w", pady=(6, 12))
        balance_slot.pack_propagate(False)

        # --- Insurance type ---
        type_frame = tk.Frame(panel, bg=t["card"])
        type_frame.pack(anchor="w", fill="x")

        styled_label(type_frame, t, "Type d'assurance", font=FONT_SMALL, fg=t["muted"]).pack(anchor="w")

        insurance_type_var = tk.StringVar(value="hospital")
        rb_hospital = tk.Radiobutton(type_frame, text="Assurance hospitalière", variable=insurance_type_var, value="hospital")
        rb_ambulatoire = tk.Radiobutton(type_frame, text="Assurance ambulatoire", variable=insurance_type_var, value="ambulatoire")

        # Store radio references for potential cleanup later
        self._insurance_radios = [rb_hospital, rb_ambulatoire]

        for rb in (rb_hospital, rb_ambulatoire):
            rb.configure(
                bg=t["card"], fg=t["text"], selectcolor=t["bg"],
                activebackground=t["card"], activeforeground=t["text"],
                font=FONT_BODY, cursor="hand2",
            )
            rb.pack(anchor="w", pady=2)

        # --- Pack selection ---
        pack_frame = tk.Frame(panel, bg=t["card"])
        pack_frame.pack(anchor="w", fill="x", pady=(12, 0))

        styled_label(pack_frame, t, "Choisir un pack", font=FONT_SMALL, fg=t["muted"]).pack(anchor="w")

        pack_listbox = tk.Listbox(pack_frame, height=6, font=FONT_BODY)
        pack_listbox.pack(anchor="w", pady=6, fill="x")

        def refresh_packs(*_args):
            pack_listbox.delete(0, tk.END)
            chosen = packs.get(insurance_type_var.get(), packs["hospital"])
            for pk_name, pk_price in chosen:
                pack_listbox.insert(tk.END, f"{pk_name} - {pk_price}€")
            try:
                pack_listbox.selection_set(0)
            except Exception:
                pass

        insurance_type_var.trace_add("write", refresh_packs)
        refresh_packs()

        # --- Price display ---
        self.price_label = styled_label(panel, t, "Prix: 200€", font=("Segoe UI Semibold", 14), fg=t["cyan"])
        self.price_label.pack(anchor="w", pady=(10, 0))

        def update_price(*_args):
            try:
                sel = pack_listbox.curselection()
                if not sel:
                    return
                text = pack_listbox.get(sel[0])
                # Extract price from text like "Pack 1 - Hospitalisation de base - 200€"
                import re
                m = re.search(r"(\d+)€", text)
                if m:
                    self.price_label.configure(text=f"Prix: {m.group(1)}€")
            except Exception:
                pass

        pack_listbox.bind("<<ListboxSelect>>", update_price)

        # --- Pay button ---
        pay_btn_slot = tk.Frame(panel, bg=t["card"], width=160, height=36)
        pay_btn_slot.pack(anchor="w", pady=(16, 0))
        pay_btn_slot.pack_propagate(False)

        # --- Insurance history ---
        history_label = styled_label(panel, t, "", font=FONT_SMALL, fg=t["muted"])
        history_label.pack(anchor="w", pady=(10, 0))

        # Store the selected customer info
        self._selected_customer_cin = None
        self._selected_customer_name = None

        def search_customer():
            cin = self._form_insurance_cin.get().strip()
            if not cin:
                messagebox.showerror("Error", "CIN requis")
                return
            if not cin.isdigit() or len(cin) != 8:
                messagebox.showerror("Error", "CIN invalide (8 chiffres)")
                return

            wb = load_workbook(FILE_PATH)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if row[1] == cin:
                    self._selected_customer_cin = str(row[1])
                    self._selected_customer_name = str(row[0])
                    self.customer_info_label.configure(
                        text=f"Client: {row[0]} | CIN: {row[1]}",
                        fg=t["green"],
                    )
                    return
            self._selected_customer_cin = None
            self._selected_customer_name = None
            self.customer_info_label.configure(text="Client non trouvé", fg=t["danger"])

        def payer():
            if not self._selected_customer_cin:
                messagebox.showerror("Error", "Veuillez d'abord rechercher un client")
                return

            try:
                balance = float(self._form_insurance_balance.get().strip().replace(",", "."))
                if balance < 0:
                    messagebox.showerror("Error", "Le solde ne peut pas être négatif")
                    return
            except ValueError:
                messagebox.showerror("Error", "Montant du solde invalide")
                return

            # Determine selected pack and price
            insurance_type = insurance_type_var.get()
            sel = pack_listbox.curselection()
            if not sel:
                messagebox.showerror("Error", "Veuillez sélectionner un pack")
                return

            text = pack_listbox.get(sel[0])
            import re
            m = re.search(r"(Pack \d+).*?(\d+)€", text)
            if not m:
                messagebox.showerror("Error", "Erreur de lecture du pack")
                return

            pack_name = m.group(1)
            price = int(m.group(2))

            if balance < price:
                messagebox.showerror(
                    "Error",
                    f"Solde insuffisant. Vous avez {balance}€ mais le pack coûte {price}€.\n"
                    "Merci de recharger votre compte.",
                )
                return

            # Process payment
            new_balance = balance - price
            insurance_label = "Hospitalière" if insurance_type == "hospital" else "Ambulatoire"

            # Save to insurance database
            add_insurance_subscription(
                cin=self._selected_customer_cin,
                customer_name=self._selected_customer_name,
                insurance_type=insurance_label,
                pack=pack_name,
                price=float(price),
                balance_before=balance,
                balance_after=new_balance,
            )

            # Update balance field
            self._form_insurance_balance.delete(0, tk.END)
            self._form_insurance_balance.insert(0, str(new_balance))

            messagebox.showinfo(
                "Success",
                f"Paiement effectué avec succès!\n"
                f"Assurance: {insurance_label}\n"
                f"Pack: {pack_name}\n"
                f"Montant débité: {price}€\n"
                f"Nouveau solde: {new_balance}€",
            )

            # Show subscription count
            subs = list_insurance_subscriptions()
            count = sum(1 for s in subs if s["cin"] == self._selected_customer_cin)
            history_label.configure(
                text=f"Ce client a {count} souscription(s) d'assurance",
                fg=t["muted"],
            )

        # Direct-on-root controls for automation
        self._form_insurance_cin = styled_entry(self.root, t)
        self._form_insurance_balance = styled_entry(self.root, t)
        self._insurance_search_btn = primary_button(self.root, t, "Search", search_customer, width=12)
        self._insurance_pay_btn = primary_button(self.root, t, "Payer", payer, width=16)
        self.form_entries = [self._form_insurance_cin, self._form_insurance_balance]

        def place_insurance_form(_e=None):
            try:
                self._place_on_slot(self._form_insurance_cin, cin_slot)
                self._place_on_slot(self._insurance_search_btn, search_btn_slot)
                self._place_on_slot(self._form_insurance_balance, balance_slot)
                self._place_on_slot(self._insurance_pay_btn, pay_btn_slot)
                for btn in self.nav_buttons.values():
                    btn.lift()
            except tk.TclError:
                pass

        self.root.after(40, place_insurance_form)

    def _view_bi(self):
        t = self.t
        host = tk.Frame(self.content, bg=t["bg"])
        host.pack(fill="both", expand=True)
        # Embed existing live Power BI panel
        from powerbi_dashboard import mount_powerbi_panel

        mount_powerbi_panel(host, theme=t)


def open_login_window():
    app = BankApp()
    app.run()


if __name__ == "__main__":
    open_login_window()
