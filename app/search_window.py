import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
from database import FILE_PATH
from ui_background import apply_window_background, style_on_background


def open_search_window():

    window = tk.Toplevel()
    window.title("Search Customer")
    window.geometry("900x600")
    apply_window_background(window)

    label_cin = tk.Label(window, text="Enter CIN")
    style_on_background(label_cin)
    label_cin.pack(pady=(40, 0))
    entry_cin = tk.Entry(window)
    style_on_background(entry_cin)
    entry_cin.pack()

    result = tk.Label(window, text="")
    style_on_background(result)
    result.pack(pady=20)

    def search():

        cin = entry_cin.get().strip()

        if not cin:
            messagebox.showerror("Error", "CIN required")
            return

        if not cin.isdigit() or len(cin) != 8:
            messagebox.showerror("Error", "Invalid CIN")
            return

        wb = load_workbook(FILE_PATH)
        ws = wb.active

        for row in ws.iter_rows(values_only=True):

            if row[1] == cin:
                result.config(
                    text=f"Name: {row[0]} | CIN: {row[1]} | Email: {row[2]}"
                )
                return

        result.config(text="Customer not found")

    btn_search = tk.Button(window, text="Search", command=search)
    style_on_background(btn_search)
    btn_search.pack()
