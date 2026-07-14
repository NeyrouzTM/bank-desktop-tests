import tkinter as tk
import re
from tkinter import messagebox
from openpyxl import load_workbook
from database import FILE_PATH, add_transfer
from ui_background import apply_window_background, style_on_background


def open_transfer_window():

    amount_pattern = re.compile(r"^\d+(\.\d{1,2})?$")

    window = tk.Toplevel()
    window.title("Transfer Money")
    window.geometry("900x600")
    apply_window_background(window)

    label_from = tk.Label(window, text="From CIN")
    style_on_background(label_from)
    label_from.pack(pady=(40, 0))
    entry_from = tk.Entry(window)
    style_on_background(entry_from)
    entry_from.pack()

    label_to = tk.Label(window, text="To CIN")
    style_on_background(label_to)
    label_to.pack()
    entry_to = tk.Entry(window)
    style_on_background(entry_to)
    entry_to.pack()

    label_amount = tk.Label(window, text="Amount (simulation)")
    style_on_background(label_amount)
    label_amount.pack()
    entry_amount = tk.Entry(window)
    style_on_background(entry_amount)
    entry_amount.pack()

    def transfer():

        from_cin = entry_from.get().strip()
        to_cin = entry_to.get().strip()
        amount = entry_amount.get().strip().replace(",", ".")

        if not from_cin or not to_cin or not amount:
            messagebox.showerror("Error", "All fields required")
            return

        if from_cin == to_cin:
            messagebox.showerror("Error", "Cannot transfer to same account")
            return

        if not from_cin.isdigit() or len(from_cin) != 8:
            messagebox.showerror("Error", "Invalid sender CIN")
            return

        if not to_cin.isdigit() or len(to_cin) != 8:
            messagebox.showerror("Error", "Invalid recipient CIN")
            return

        if not amount_pattern.fullmatch(amount):
            messagebox.showerror("Error", "Amount format is invalid")
            return

        try:
            amount = float(amount)
            if amount <= 0:
                messagebox.showerror("Error", "Invalid amount")
                return
        except ValueError:
            messagebox.showerror("Error", "Amount must be numeric")
            return

        wb = load_workbook(FILE_PATH)
        ws = wb.active

        from_exists = False
        to_exists = False

        for row in ws.iter_rows(values_only=True):

            if row[1] == from_cin:
                from_exists = True

            if row[1] == to_cin:
                to_exists = True

        if not from_exists or not to_exists:
            messagebox.showerror("Error", "Invalid CIN(s)")
            return

        add_transfer(from_cin, to_cin, amount, status="Success")

        messagebox.showinfo("Success", "Transfer simulated successfully")

        entry_from.delete(0, tk.END)
        entry_to.delete(0, tk.END)
        entry_amount.delete(0, tk.END)

    btn_transfer = tk.Button(window, text="Transfer", command=transfer)
    style_on_background(btn_transfer)
    btn_transfer.pack(pady=10)
